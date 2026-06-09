"""사주도령 데일리 트래커(.xlsx) 생성기 — 플랜 §12.

빈 템플릿을 생성한다(시트 5개). openpyxl 필요(전역 파이썬에 설치됨).
재생성 시 기존 파일을 덮어쓰므로, 데이터가 쌓인 뒤에는 백업 후 실행할 것.

사용: python saju-growth-system/automation/tracker_build.py [출력경로]
기본 출력: saju-growth-system/tracking/daily-tracker.xlsx
"""

import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEAD_FILL = PatternFill("solid", fgColor="2F3B30")
HEAD_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="EAF0E6")
TARGET = 200000


def _style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build(path):
    wb = Workbook()

    # 1) 안내
    ws = wb.active
    ws.title = "안내"
    guide = [
        ["사주도령 데일리 트래커 — 사용 안내", ""],
        ["", ""],
        ["목적", "측정→교정 루프. 매일 콘텐츠로그·판매로그 입력 → 일일요약이 자동 집계."],
        [
            "핵심 KPI",
            "① 외부링크 클릭수(현재 기준선: IG 90일 66건) ② 문의→결제 전환율 ③ 일매출(목표 200,000)",
        ],
        [
            "재활성 귀속",
            "카카오 5일 캠페인 유입은 신청 시 '재오픈' 키워드로 구분 → 판매로그 first_keyword에 기록",
        ],
        ["시트", "콘텐츠로그 / 주제별성과 / 판매로그 / 일일요약 / 캠페인_재활성"],
        [
            "주제별 측정",
            "콘텐츠로그의 topic 열에 daily_hooks.py의 topic_id를 적으면 주제별성과 시트가 조회·참여·문의를 자동 롤업 → 1위 주제 더블다운(릴스·Shorts 확산).",
        ],
        [
            "개인정보",
            "판매로그 이름은 익명코드(예: K-001) 사용. 생년월일시 등 민감정보는 여기 적지 말 것.",
        ],
        ["가격", "입문 9,900 / 코어 29,000 / 프리미엄 59,000 (옵션 궁합 +20,000·당일 +5,000)"],
        ["주의", "재생성 스크립트는 덮어쓰므로 데이터 입력 후엔 백업 후 실행."],
    ]
    for r in guide:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)
    _widths(ws, [16, 110])
    for row in range(3, len(guide) + 1):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)

    # 2) 콘텐츠로그 (플랜 §12 필드)
    ws = wb.create_sheet("콘텐츠로그")
    cols = [
        "date",
        "channel",
        "topic",  # 주제 식별자(daily_hooks.py의 topic_id) — 주제별성과 롤업의 키
        "content_id",
        "hook_type",
        "url",
        "cta",
        "link_variant",
        "kakao_keyword",
        "views",
        "comments",
        "saves",
        "shares",
        "profile_clicks",
        "ext_link_clicks",
        "kakao_inflow",
        "menu_clicks",
        "inquiries",
        "payments",
        "revenue",
        "offer",
        "pdf_delivered",
        "extra_q",
        "repurchase",
        "notes",
    ]
    ws.append(cols)
    _style_header(ws, len(cols))
    _widths(
        ws,
        [11, 11, 16, 14, 14, 22, 18, 12, 12, 8, 9, 7, 7, 12, 13, 11, 11, 9, 9, 10, 12, 11, 8, 10, 26],
    )

    def _col(name):
        return get_column_letter(cols.index(name) + 1)

    # 드롭다운(컬럼 위치는 cols 인덱스로 계산 — topic 삽입에도 안전)
    dv_ch = DataValidation(
        type="list", formula1='"threads,instagram,kakao,youtube,tiktok"', allow_blank=True
    )
    dv_offer = DataValidation(
        type="list", formula1='"입문9900,코어29000,프리미엄59000,궁합옵션,-"', allow_blank=True
    )
    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    ws.add_data_validation(dv_ch)
    dv_ch.add(f"{_col('channel')}2:{_col('channel')}2000")
    ws.add_data_validation(dv_offer)
    dv_offer.add(f"{_col('offer')}2:{_col('offer')}2000")
    ws.add_data_validation(dv_yn)
    dv_yn.add(f"{_col('pdf_delivered')}2:{_col('pdf_delivered')}2000")
    dv_yn.add(f"{_col('repurchase')}2:{_col('repurchase')}2000")

    # 2b) 주제별성과 (콘텐츠로그에서 topic 기준 자동 롤업 — 측정→교정 루프)
    wst = wb.create_sheet("주제별성과")
    tcols = [
        "topic",
        "게시수(자동)",
        "총조회(자동)",
        "댓글(자동)",
        "공유(자동)",
        "외부링크(자동)",
        "문의(자동)",
        "결제(자동)",
        "조회당참여(자동)",
    ]
    wst.append(tcols)
    _style_header(wst, len(tcols))
    _widths(wst, [20, 12, 12, 10, 9, 12, 10, 9, 14])
    CL = "콘텐츠로그"
    c_topic, c_views = _col("topic"), _col("views")
    c_comments, c_shares = _col("comments"), _col("shares")
    c_ext, c_inq, c_pay = _col("ext_link_clicks"), _col("inquiries"), _col("payments")
    # 알려진 topic_id 시드(daily_hooks.py TOPIC_BANK와 일치) + 신규용 빈 행
    seed_topics = [
        "money-no-keep",
        "move-or-wait",
        "effort-no-result",
        "relationship-loop",
        "luck-turning-signal",
        "career-fit",
        "decision-delayed",
        "same-saju-diff",
    ]
    for r, tid in enumerate(seed_topics + [""] * 6, start=2):
        wst.cell(row=r, column=1).value = tid
        wst.cell(row=r, column=2).value = f"=COUNTIF({CL}!${c_topic}:${c_topic},$A{r})"
        wst.cell(row=r, column=3).value = f"=SUMIF({CL}!${c_topic}:${c_topic},$A{r},{CL}!${c_views}:${c_views})"
        wst.cell(row=r, column=4).value = f"=SUMIF({CL}!${c_topic}:${c_topic},$A{r},{CL}!${c_comments}:${c_comments})"
        wst.cell(row=r, column=5).value = f"=SUMIF({CL}!${c_topic}:${c_topic},$A{r},{CL}!${c_shares}:${c_shares})"
        wst.cell(row=r, column=6).value = f"=SUMIF({CL}!${c_topic}:${c_topic},$A{r},{CL}!${c_ext}:${c_ext})"
        wst.cell(row=r, column=7).value = f"=SUMIF({CL}!${c_topic}:${c_topic},$A{r},{CL}!${c_inq}:${c_inq})"
        wst.cell(row=r, column=8).value = f"=SUMIF({CL}!${c_topic}:${c_topic},$A{r},{CL}!${c_pay}:${c_pay})"
        wst.cell(row=r, column=9).value = f'=IF(C{r}=0,"",(D{r}+E{r})/C{r})'
    wst["A16"].value = "↑ 조회 내림차순 정렬 후 1위 주제를 릴스·Shorts로 확산(repurpose.py). 신규 주제는 빈 행에 topic_id 추가."
    wst["A16"].fill = NOTE_FILL

    # 3) 판매로그
    ws = wb.create_sheet("판매로그")
    cols = [
        "date",
        "익명코드",
        "offer",
        "price",
        "source_channel",
        "first_keyword",
        "payment_status",
        "delivery_status",
        "refund",
        "notes",
    ]
    ws.append(cols)
    _style_header(ws, len(cols))
    _widths(ws, [11, 10, 14, 10, 14, 14, 13, 13, 9, 30])
    dv_offer2 = DataValidation(
        type="list", formula1='"입문9900,코어29000,프리미엄59000,궁합옵션"', allow_blank=True
    )
    dv_src = DataValidation(
        type="list", formula1='"threads,instagram,kakao_재오픈,kakao_기타,기타"', allow_blank=True
    )
    dv_pay = DataValidation(type="list", formula1='"대기,입금완료,취소,환불"', allow_blank=True)
    dv_del = DataValidation(type="list", formula1='"대기,발송완료"', allow_blank=True)
    ws.add_data_validation(dv_offer2)
    dv_offer2.add("C2:C2000")
    ws.add_data_validation(dv_src)
    dv_src.add("E2:E2000")
    ws.add_data_validation(dv_pay)
    dv_pay.add("G2:G2000")
    ws.add_data_validation(dv_del)
    dv_del.add("H2:H2000")

    # 4) 일일요약 (판매로그에서 자동 집계)
    ws = wb.create_sheet("일일요약")
    cols = [
        "date",
        "total_views",
        "ext_link_clicks",
        "kakao_inflow",
        "inquiries",
        "결제건수(자동)",
        "매출(자동)",
        "목표",
        "갭(자동)",
    ]
    ws.append(cols)
    _style_header(ws, len(cols))
    _widths(ws, [12, 12, 14, 12, 10, 14, 14, 10, 14])
    for i in range(2, 22):  # 20일치 빈 행 + 수식
        ws.cell(
            row=i, column=6
        ).value = f'=COUNTIFS(판매로그!$A:$A,$A{i},판매로그!$G:$G,"입금완료")'
        ws.cell(
            row=i, column=7
        ).value = f'=SUMIFS(판매로그!$D:$D,판매로그!$A:$A,$A{i},판매로그!$G:$G,"입금완료")'
        ws.cell(row=i, column=8).value = TARGET
        ws.cell(row=i, column=9).value = f"=H{i}-G{i}"

    # 5) 캠페인_재활성 (5일)
    ws = wb.create_sheet("캠페인_재활성")
    cols = [
        "Day",
        "메시지",
        "형태",
        "발송시각",
        "비용",
        "도달",
        "클릭",
        "문의(재오픈)",
        "결제",
        "매출",
        "notes",
    ]
    ws.append(cols)
    _style_header(ws, len(cols))
    _widths(ws, [6, 22, 16, 12, 9, 9, 9, 12, 8, 10, 24])
    seed = [
        ["D1", "M1 가치+재오픈 예고", "유료 채널메시지", "", 70000, "", "", "", "", "", ""],
        ["D2", "M2 스토리/오픈루프", "무료 소식", "", 0, "", "", "", "", "", ""],
        ["D3", "M3 오퍼 공개(핵심)", "유료 채널메시지", "", 70000, "", "", "", "", "", ""],
        ["D4", "M4 히든베니핏+후기", "무료 소식", "", 0, "", "", "", "", "", ""],
        ["D5", "M5 마감 D-day", "유료 채널메시지", "", 70000, "", "", "", "", "", ""],
    ]
    for r in seed:
        ws.append(r)
    ws.cell(row=8, column=1).value = "합계"
    ws.cell(row=8, column=5).value = "=SUM(E2:E6)"
    ws.cell(row=8, column=10).value = "=SUM(J2:J6)"
    ws.cell(row=9, column=1).value = "ROI"
    ws.cell(row=9, column=2).value = '=IF(E8=0,"",J8/E8)'

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"[생성] {path}")
    print(f"[시트] {wb.sheetnames}")


def main():
    default = os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "tracking",
        "daily-tracker.xlsx",
    )
    path = sys.argv[1] if len(sys.argv) > 1 else default
    build(path)


if __name__ == "__main__":
    main()
