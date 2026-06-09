#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""사주도령 7일 결산 — 트래커에서 승자 루트/오퍼/훅 판정. 플랜 §10 Day7.

daily-tracker.xlsx(콘텐츠로그·판매로그)를 읽어 누적을 4축으로 집계한다.
 - 훅(hook_type)별: 도달·외부링크클릭·문의·결제·매출 → 승자 훅
 - 채널별: 동일 → 승자 채널
 - 오퍼별: 결제건수·매출 → 승자 오퍼(입문/코어/프리미엄/궁합)
 - 루트별: first_keyword '재오픈'(웜·Route D 채팅직행) vs '소식'(콜드·Route B 소식랜딩) → 승자 루트
 - 총매출 vs 목표(일 200,000 × days)

외부 API 의존 없음(openpyxl 파일 읽기만). 데이터가 비어 있으면 안내만 출력.
판매로그 payment_status='입금완료'인 행만 매출/결제로 인정(대기·취소·환불 제외).

사용:
  python weekly_review.py --xlsx ../tracking/daily-tracker.xlsx
  python weekly_review.py --xlsx daily-tracker.xlsx --out day7-summary.md --days 7
의존: openpyxl(전역 파이썬에 설치됨).
"""

import argparse
import os
import sys
from collections import defaultdict

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("openpyxl이 필요합니다: pip install openpyxl\n")
    sys.exit(2)

TARGET_PER_DAY = 200000


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _rows(ws):
    """워크시트를 헤더 기반 dict 리스트로. 1행=헤더."""
    it = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(it)]
    except StopIteration:
        return []
    out = []
    for row in it:
        if row is None or all(c is None for c in row):
            continue
        out.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    return out


def _route(keyword):
    """first_keyword/kakao_keyword → 루트 라벨."""
    k = (str(keyword) if keyword is not None else "").strip()
    if "재오픈" in k:
        return "Route D(웜·채팅직행)"
    if "소식" in k:
        return "Route B(콜드·소식랜딩)"
    if "팔로업" in k:
        return "팔로업(2차)"
    return "기타/미상"


def _winner(d, key):
    """집계 dict에서 key값 최대 항목명."""
    best = None
    for name, agg in d.items():
        v = agg.get(key, 0)
        if best is None or v > best[1]:
            best = (name, v)
    return best[0] if best else "-"


def collect(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    content = _rows(wb["콘텐츠로그"]) if "콘텐츠로그" in wb.sheetnames else []
    sales = _rows(wb["판매로그"]) if "판매로그" in wb.sheetnames else []
    wb.close()

    by_hook = defaultdict(lambda: defaultdict(float))
    by_channel = defaultdict(lambda: defaultdict(float))
    for r in content:
        hook = (str(r.get("hook_type")) if r.get("hook_type") else "미상").strip()
        ch = (str(r.get("channel")) if r.get("channel") else "미상").strip()
        for grp, name in ((by_hook, hook), (by_channel, ch)):
            g = grp[name]
            g["n"] += 1
            g["views"] += _num(r.get("views"))
            g["ext_clicks"] += _num(r.get("ext_link_clicks"))
            g["inquiries"] += _num(r.get("inquiries"))
            g["payments"] += _num(r.get("payments"))
            g["revenue"] += _num(r.get("revenue"))

    by_offer = defaultdict(lambda: defaultdict(float))
    by_route = defaultdict(lambda: defaultdict(float))
    total_rev = 0.0
    total_paid = 0
    for r in sales:
        status = (str(r.get("payment_status")) if r.get("payment_status") else "").strip()
        if status != "입금완료":
            continue
        price = _num(r.get("price"))
        offer = (str(r.get("offer")) if r.get("offer") else "미상").strip()
        route = _route(r.get("first_keyword"))
        by_offer[offer]["paid"] += 1
        by_offer[offer]["revenue"] += price
        by_route[route]["paid"] += 1
        by_route[route]["revenue"] += price
        total_rev += price
        total_paid += 1

    return {
        "content_n": len(content),
        "sales_n": len(sales),
        "by_hook": by_hook,
        "by_channel": by_channel,
        "by_offer": by_offer,
        "by_route": by_route,
        "total_rev": total_rev,
        "total_paid": total_paid,
    }


def render(xlsx_path, days=7):
    d = collect(xlsx_path)
    out = []
    out.append(f"# 7일 결산 — 승자 판정 ({os.path.basename(xlsx_path)})")
    out.append("")

    if d["content_n"] == 0 and d["sales_n"] == 0:
        out.append("> 아직 입력된 데이터가 없습니다. 캠페인 진행 중/후 콘텐츠로그·판매로그를")
        out.append("> 채운 뒤 다시 실행하세요. (판매로그는 payment_status='입금완료'만 매출 인정)")
        return "\n".join(out)

    target = TARGET_PER_DAY * days
    rev = d["total_rev"]
    out.append(f"- 기간: {days}일 / 누적 결제 {d['total_paid']}건 / 누적 매출 **{int(rev):,}원**")
    out.append(
        f"- 목표({TARGET_PER_DAY:,}×{days}일 = {target:,}원) 대비 {rev / target * 100:.0f}%"
        f" / 일평균 {int(rev / days):,}원"
    )
    out.append("")

    out.append("## 1) 훅(hook_type)별 — 승자 훅 = 외부링크클릭·문의 기준")
    out.append("| 훅 | 게시 | 도달 | 외부클릭 | 문의 | 결제 | 매출 |")
    out.append("|---|---:|---:|---:|---:|---:|---:|")
    for name, g in sorted(d["by_hook"].items(), key=lambda x: -x[1]["ext_clicks"]):
        out.append(
            f"| {name} | {int(g['n'])} | {int(g['views']):,} | {int(g['ext_clicks']):,}"
            f" | {int(g['inquiries'])} | {int(g['payments'])} | {int(g['revenue']):,} |"
        )
    out.append(
        f"\n**승자 훅(외부클릭)**: `{_winner(d['by_hook'], 'ext_clicks')}` → 다음 주 더블다운."
    )
    out.append("")

    out.append("## 2) 채널별 — 승자 채널")
    out.append("| 채널 | 게시 | 도달 | 외부클릭 | 문의 | 결제 | 매출 |")
    out.append("|---|---:|---:|---:|---:|---:|---:|")
    for name, g in sorted(d["by_channel"].items(), key=lambda x: -x[1]["ext_clicks"]):
        out.append(
            f"| {name} | {int(g['n'])} | {int(g['views']):,} | {int(g['ext_clicks']):,}"
            f" | {int(g['inquiries'])} | {int(g['payments'])} | {int(g['revenue']):,} |"
        )
    out.append(f"\n**승자 채널(외부클릭)**: `{_winner(d['by_channel'], 'ext_clicks')}`")
    out.append("")

    out.append("## 3) 오퍼별 — 승자 오퍼(매출 기여)")
    out.append("| 오퍼 | 결제건수 | 매출 |")
    out.append("|---|---:|---:|")
    for name, g in sorted(d["by_offer"].items(), key=lambda x: -x[1]["revenue"]):
        out.append(f"| {name} | {int(g['paid'])} | {int(g['revenue']):,} |")
    out.append(f"\n**승자 오퍼(매출)**: `{_winner(d['by_offer'], 'revenue')}`")
    out.append("")

    out.append("## 4) 루트별 — Route D(웜·채팅직행) vs Route B(콜드·소식랜딩)")
    out.append("| 루트 | 결제건수 | 매출 |")
    out.append("|---|---:|---:|")
    for name, g in sorted(d["by_route"].items(), key=lambda x: -x[1]["revenue"]):
        out.append(f"| {name} | {int(g['paid'])} | {int(g['revenue']):,} |")
    out.append(
        f"\n**승자 루트(매출)**: `{_winner(d['by_route'], 'revenue')}`"
        " → 다음 주 이 동선에 자원 집중."
    )
    out.append("")

    out.append("## 다음 주 결정(3개만 — day7-decision-template.md에 기록)")
    out.append(
        f"- [ ] 더블다운: 승자 훅 `{_winner(d['by_hook'], 'ext_clicks')}`"
        f" + 승자 채널 `{_winner(d['by_channel'], 'ext_clicks')}`"
    )
    out.append(f"- [ ] 자원집중: 승자 루트 `{_winner(d['by_route'], 'revenue')}`")
    out.append("- [ ] 중단: 외부클릭·전환 최하위 훅/채널 1개")
    return "\n".join(out)


def main(argv=None):
    p = argparse.ArgumentParser(description="사주도령 7일 결산(승자 판정)")
    p.add_argument("--xlsx", required=True, help="daily-tracker.xlsx 경로")
    p.add_argument("--out", default=None, help="출력 md(미지정 시 표준출력)")
    p.add_argument("--days", type=int, default=7, help="결산 일수(기본 7)")
    args = p.parse_args(argv)

    # Windows cp949 콘솔에서 유니코드(— 등) 출력 깨짐 방지
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    if not os.path.exists(args.xlsx):
        sys.stderr.write(f"파일 없음: {args.xlsx}\n")
        return 2

    text = render(args.xlsx, args.days)
    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(text + "\n")
        print("생성: " + args.out)
    else:
        print(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
